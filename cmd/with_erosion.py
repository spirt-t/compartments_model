import numpy as np
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy.integrate import odeint
import pandas as pd
import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
from scipy.optimize import curve_fit


def check():
    try:
        global geometry
        if geometry_var.get() == 1:
            geometry = 'slab'
        if geometry_var.get() == 2:
            geometry = 'cylinder'
        if geometry_var.get() == 3:
            geometry = 'sphere'
        global plot
        plot = []
        if plot_q.get() == 1:
            plot.append('Q')
        if plot_flux.get() == 1:
            plot.append('flux')
        if plot_total_flux.get() == 1:
            plot.append('total_flux')
        if plot_concentration.get() == 1:
            plot.append('concentration')
        if len(plot) > 0:
            make_graphs.set(1)

        global define_parameters
        define_parameters = True
        global choose_geometry
        choose_geometry = True
        global define_functions
        define_functions = True
        global Input_Parameters
        Input_Parameters = define_input_parameters()
        if len(Input_Parameters) < 3:
            raise ValueError
        global Output_Parameters
        Output_Parameters = define_output_parameters()
        if len(Output_Parameters) < 3:
            raise ValueError
        global N
        N = get_entry_value(entry_N)  # Number of compartments
        N = int(N)
        global count
        count = int(get_entry_value(entry_count))  # Number of particles / fibres
        global mass
        mass = get_entry_value(entry_mass)  # Mass of drug loaded (M)
        global mass_bound
        mass_bound = get_entry_value(entry_mass_bound)  # Mass of drug bound to polymer (M)
        global v
        v = get_entry_value(entry_v)  # Erosion front velocity (L/T)
        global D
        D = get_entry_value(entry_D)  # Diffusion coefficient (L^2/T)
        global R0
        R0 = get_entry_value(entry_R0)  # Initial radius/width (L)
        global h
        h = get_entry_value(entry_h)  # Height of cylinder (L)
        global slab_area
        slab_area = get_entry_value(entry_slab_area)  # Area if solving in a slab (L^2)
        global kp
        kp = N * D / h  # Permeability coefficient with constant volume (L/T)
        global t0
        t0 = get_entry_value(entry_t0)  # Start time (T)
        global tm
        tm = get_entry_value(entry_tm)  # End time (T)
        global ti
        ti = get_entry_value(entry_ti)  # How many time intervals
        ti = int(ti)
        global tolerance
        tolerance = get_entry_value(entry_tolerance)  # Determine when to stop kpn2 from growing
        global SD
        SD = get_entry_value(entry_SD)  # Standard deviation of distribution (L)
        global bins
        bins = int(get_entry_value(entry_bins))  # Number of bins in distribution
        # This next part just converts units
        if Input_Parameters[0] != Output_Parameters[0]:
            if Input_Parameters[0] == 'pg':
                mass = mass * 1e-12
                mass_bound = mass_bound * 1e-12
            if Input_Parameters[0] == 'ng':
                mass = mass * 1e-9
                mass_bound = mass_bound * 1e-9
            if Input_Parameters[0] == 'ug':
                mass = mass * 1e-6
                mass_bound = mass_bound * 1e-6
            if Input_Parameters[0] == 'mg':
                mass = mass * 1e-3
                mass_bound = mass_bound * 1e-3
            if Input_Parameters[0] == 'kg':
                mass = mass * 1e3
                mass_bound = mass_bound * 1e3
            if Output_Parameters[0] == 'pg':
                mass = mass * 1e12
                mass_bound = mass_bound * 1e12
            if Output_Parameters[0] == 'ng':
                mass = mass * 1e9
                mass_bound = mass_bound * 1e9
            if Output_Parameters[0] == 'ug':
                mass = mass * 1e6
                mass_bound = mass_bound * 1e6
            if Output_Parameters[0] == 'mg':
                mass = mass * 1e3
                mass_bound = mass_bound * 103
            if Output_Parameters[0] == 'kg':
                mass = mass * 1e-3
                mass_bound = mass_bound * 10 - 3

        if Input_Parameters[1] != Output_Parameters[1]:
            if Input_Parameters[1] == 'pm':
                v = v * 1e-12
                D = D * 1e-24
                R0 = R0 * 1e-12
                h = h * 1e-12
                slab_area = slab_area * 1e-12
                kp = kp * 1e-12
                SD = SD * 1e-12
            if Input_Parameters[1] == 'nm':
                v = v * 1e-9
                D = D * 1e-18
                R0 = R0 * 1e-9
                h = h * 1e-9
                slab_area = slab_area * 1e-18
                kp = kp * 1e-9
                SD = SD * 1e-9
            if Input_Parameters[1] == 'um':
                v = v * 1e-6
                D = D * 1e-12
                R0 = R0 * 1e-6
                h = h * 1e-6
                slab_area = slab_area * 1e-12
                kp = kp * 1e-6
                SD = SD * 1e-6
            if Input_Parameters[1] == 'mm':
                v = v * 1e-3
                D = D * 1e-6
                R0 = R0 * 1e-3
                h = h * 1e-3
                slab_area = slab_area * 1e-6
                kp = kp * 1e-3
                SD = SD * 1e-3
            if Input_Parameters[1] == 'cm':
                v = v * 1e-2
                D = D * 1e-4
                R0 = R0 * 1e-2
                h = h * 1e-2
                slab_area = slab_area * 1e-4
                kp = kp * 1e-2
                SD = SD * 1e-2
            if Output_Parameters[1] == 'pm':
                v = v * 1e12
                D = D * 1e24
                R0 = R0 * 1e12
                h = h * 1e12
                slab_area = slab_area * 1e12
                kp = kp * 1e12
                SD = SD * 1e12
            if Output_Parameters[1] == 'nm':
                v = v * 1e9
                D = D * 1e18
                R0 = R0 * 1e9
                h = h * 1e9
                slab_area = slab_area * 1e18
                kp = kp * 1e9
                SD = SD * 1e9
            if Output_Parameters[1] == 'um':
                v = v * 1e6
                D = D * 1e12
                R0 = R0 * 1e6
                h = h * 1e6
                slab_area = slab_area * 1e12
                kp = kp * 1e6
                SD = SD * 1e6
            if Output_Parameters[1] == 'mm':
                v = v * 1e3
                D = D * 1e6
                R0 = R0 * 1e3
                h = h * 1e3
                slab_area = slab_area * 1e6
                kp = kp * 1e3
                SD = SD * 1e3
            if Output_Parameters[1] == 'cm':
                v = v * 1e2
                D = D * 1e4
                R0 = R0 * 1e2
                h = h * 1e2
                slab_area = slab_area * 1e4
                kp = kp * 1e2
                SD = SD * 1e2

        if Input_Parameters[2] != Output_Parameters[2]:
            if Input_Parameters[2] == 'min':
                v = v / 60
                D = D / 60
                kp = kp / 60
                t0 = t0 * 60
                tm = tm * 60
            if Input_Parameters[2] == 'hour':
                v = v / 3600
                D = D / 3600
                kp = kp / 3600
                t0 = t0 * 3600
                tm = tm * 3600
            if Input_Parameters[2] == 'day':
                v = v / 86400
                D = D / 86400
                kp = kp / 86400
                t0 = t0 * 86400
                tm = tm * 86400
            if Input_Parameters[2] == 'year':
                v = v / 31536000
                D = D / 31536000
                kp = kp / 31536000
                t0 = t0 * 31536000
                tm = tm * 31536000
            if Output_Parameters[2] == 'min':
                v = v * 60
                D = D * 60
                kp = kp * 60
                t0 = t0 / 60
                tm = tm / 60
            if Output_Parameters[2] == 'hour':
                v = v * 3600
                D = D * 3600
                kp = kp * 3600
                t0 = t0 / 3600
                tm = tm / 3600
            if Output_Parameters[2] == 'day':
                v = v * 86400
                D = D * 86400
                kp = kp * 86400
                t0 = t0 / 86400
                tm = tm / 86400
            if Output_Parameters[2] == 'year':
                v = v * 31536000
                D = D * 31536000
                kp = kp * 31536000
                t0 = t0 / 31536000
                tm = tm / 31536000

        if Output_Parameters[0] == 'ug':
            Output_Parameters[0] = '\mu g'  # Fixes display of mu in figure titles

        if Output_Parameters[1] == 'um':
            Output_Parameters[1] = '\mu m'  # Fixes display of mu in figure titles

        global experimental_time_column_number
        if entry_time_column.get() != '':
            experimental_time_column_number = int(get_entry_value(entry_time_column))
        else:
            experimental_time_column_number = -1

        global experimental_flux_column_number
        if entry_flux_column.get() != '':
            experimental_flux_column_number = int(get_entry_value(entry_flux_column))
        else:
            experimental_flux_column_number = -1

        global experimental_total_flux_column_number
        if entry_total_flux_column.get() != '':
            experimental_total_flux_column_number = int(get_entry_value(entry_total_flux_column))
        else:
            experimental_total_flux_column_number = -1

        global experimental_release_column_number
        if entry_release_column.get() != '':
            experimental_release_column_number = int(get_entry_value(entry_release_column))
        else:
            experimental_release_column_number = -1

        global experimental_x_column_number  # This is desplayed wrong in GUI heading
        if entry_x_column.get() != '':
            experimental_x_column_number = int(get_entry_value(entry_x_column))
        else:
            experimental_x_column_number = -1

        global experimental_concentration_column_number  # This is desplayed wrong in GUI heading
        if entry_concentration_column.get() != '':
            experimental_concentration_column_number = int(get_entry_value(entry_concentration_column))
        else:
            experimental_concentration_column_number = -1

    except ValueError:
        tk.messagebox.showerror("Error", "All parameters must be entered!")
    except NameError:
        tk.messagebox.showerror("Error", "All parameters must be entered!")

    # print("check end")


def define_params():
    global experimental_time, experimental_flux, y_regression, experimental_total_flux
    global experimental_release, experimental_x, experimental_concentration
    global t
    t = np.linspace(t0, tm, ti)  # Setup array with time values

    if enter_experimental_data.get() is True:

        if experimental_time_column_number > 0:
            experimental_time = [
                imported_data_1[experimental_time_column_number - 1, :],
            ]

        if experimental_flux_column_number > 0:
            experimental_flux = [
                imported_data_1[experimental_flux_column_number - 1, :],
            ]
            y_regression = experimental_flux
        else:
            experimental_flux = []

        if experimental_total_flux_column_number > 0:
            experimental_total_flux = [
                imported_data_1[experimental_total_flux_column_number - 1, :],
            ]
            y_regression = experimental_total_flux
        else:
            experimental_total_flux = []

        if experimental_release_column_number > 0:
            experimental_release = [
                imported_data_1[experimental_release_column_number - 1, :],
            ]
            y_regression = experimental_release
        else:
            experimental_release = []

        if experimental_x_column_number > 0:
            experimental_x = [
                imported_data_1[experimental_x_column_number - 1, :],
            ]
        else:
            experimental_x = []

        if experimental_concentration_column_number > 0:
            experimental_concentration = [
                imported_data_1[experimental_concentration_column_number, :],
            ]
            y_regression = experimental_concentration
        else:
            experimental_concentration = []


def calculate(time, D, v):
    global R0
    global c
    global J_out, J_total, t_deg, solver
    # print(enter_experimental_data.get())
    # print(use_distribution.get())
    # print(make_graphs.get())

    t = np.linspace(t0, tm, ti)  # Setup array with time values

    global Q1, enter_experimental_data
    Q1 = np.zeros(ti)  # Setup array that will hold values for release
    J_out = np.zeros(len(t))  # Setup array that will hold flux values
    J_total = np.zeros(len(t))  # Setup array that will hold total flux values

    "###-=-=-=-=-=-=-=-=- Define all the Functions -=-=-=-=-=-=-=-=-###"
    if define_functions is True:

        def SLAB(c, t):
            """ Setup the system of n equations"""
            # kp = permeability coefficient (um/hour)
            # kpn = permeability coefficient for the n^th compartment
            kp = N * D / R0
            kpn = 2 * D / ((N - n + 2) * R0 / N - v * t)
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)
            dc = []
            dc.append(kp * N / R0 * (-c[0] + c[1]))
            for i in range(1, n - 2):
                dc.append(kp * N / R0 * (c[i - 1] - 2 * c[i] + c[i + 1]))
            dc.append(
                kp * N / R0 * (c[n - 3] - c[n - 2])
                + kpn * N / R0 * (c[n - 1] - c[n - 2])
            )
            dc.append(1 / (R0 - v * t - R0 * (n - 1) / N)
                      * (kpn * (c[n - 2] - c[n - 1]) - kpn2 * c[n - 1])
                      )
            return np.array(dc)

        def two_SLAB(c, t):
            """ This system is solved if two compartments remain"""
            # kpn = permeability coefficient (um/hour)
            dc = []
            kpn = 2 * D / (((N - n + 2) * R0 / N - v * t))
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)
            dc.append(kpn * N / R0 * (-c[0] + c[1]))
            dc.append(1 / (R0 - v * t - R0 * (n - 1) / N) * (kpn * (-c[1] + c[0]) - kpn2 * c[1]))
            return np.array(dc)

        def single_SLAB(c, t):
            """This single ODE is solved if only one compartment remains"""
            # kpn = permeability coefficient (um/hour)
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)
            dc = -kpn2 * c[n - 1] / (R0 - v * t)
            return np.array(dc)

        def CYLINDER(c, t):
            """ Setup the system of n equations"""
            # kp = permeability coefficient (um/hour)
            # kpn = permeability coefficient for the n^th compartment
            kp = N * D / R0
            kpn = 2 * D / ((N - n + 2) * R0 / N - v * t)
            kpn2 = 2 * D / ((N - n + 1) * R0 / N - v * t)
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)

            dc = []
            dc.append(2 * kp * N / R0 * (c[1] - c[0]))
            for i in range(1, n - 2):
                dc.append(
                    2 * kp * N / (R0 * ((i + 1) ** 2 - (i) ** 2))
                    * ((i) * (c[i - 1] - c[i]) + (i + 1) * (c[i + 1] - c[i]))
                )
            dc.append(
                2 * N / (R0 * ((n - 1) ** 2 - (n - 2) ** 2))
                * (kp * (n - 2) * (c[n - 3] - c[n - 2])
                   + kpn * (n - 1) * (c[n - 1] - c[n - 2]))
            )
            dc.append(
                2 / ((R0 - v * t) ** 2 - (R0 * (n - 1) / N) ** 2)
                * (kpn * R0 * (n - 1) / n * (c[n - 2] - c[n - 1])
                   - kpn2 * (R0 - v * t) * c[n - 1])
            )
            return np.array(dc)

        def two_CYLINDER(c, t):
            """ This system is solved if two compartments remain"""
            # kpn = permeability coefficient (um/hour)
            dc = []
            kpn = 2 * D / ((N - n + 2) * R0 / N - v * t)
            kpn2 = 2 * D / ((N - n + 1) * R0 / N - v * t)
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)
            dc.append(2 * kpn * N / R0 * (c[1] - c[0]))
            dc.append(2 / ((R0 - v * t) ** 2 - (R0 / N) ** 2)
                      * (kpn * R0 / 2 * (c[0] - c[1]) - kpn2 * (R0 - v * t) * c[1]))
            return np.array(dc)

        def single_CYLINDER(c, t):
            """This single ODE is solved if only one compartment remains"""
            # kpn = permeability coefficient (um/hour)
            kpn2 = 2 * D / ((N - n + 1) * R0 / N - v * t)
            if kpn2 >= 2 * D / (tolerance * R0 / N):
                kpn2 = 2 * D / (tolerance * R0 / N)
            dc = -2 * kpn2 / (R0 - v * t) * c[0]
            return np.array(dc)

        def SPHERE(c, t):
            """ Setup the system of n equations"""
            # kp = permeability coefficient (um/hour)
            # kpn = permeability coefficient for the n^th compartment
            kp = N * D / R0
            kpn = 2 * D / ((N - n + 2) * R0 / N - v * t)
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * (N - n + 1) * R0 / N):
                kpn2 = 2 * D / (tolerance * (N - n + 1) * R0 / N)
            dc = []
            dc.append(3 * kp * N / R0 * (c[1] - c[0]))
            for ii in range(1, n - 2):
                i = ii + 1
                dc.append(
                    3 * kp * N * i ** 2 / (R0 * (i ** 3 - (i - 1) ** 3))
                    * ((i - 1) ** 2 / i ** 2 * (c[ii - 1] - c[ii]) - c[ii] + c[ii + 1])
                )
            dc.append(
                3 * N / (R0 * ((n - 1) ** 3 - (n - 2) ** 3))
                * (kp * (n - 2) ** 2 * (c[n - 3] - c[n - 2])
                   + kpn * (n - 1) ** 2 * (c[n - 1] - c[n - 2]))
            )
            dc.append(
                3 / ((R0 - v * t) ** 3 - (R0 * (n - 1) / N) ** 3)
                * (kpn * (R0 * (n - 1) / N) ** 2 * (c[n - 2] - c[n - 1])
                   - kpn2 * (R0 - v * t) ** 2 * c[n - 1])
            )
            return np.array(dc)

        def two_SPHERE(c, t):
            """ This system is solved if two compartments remain"""
            # kpn = permeability coefficient (um/hour)
            dc = []
            kpn = 2 * D / ((N - n + 2) * R0 / N - v * t)
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * (N - n + 1) * R0 / N):
                kpn2 = 2 * D / (tolerance * (N - n + 1) * R0 / N)
            dc.append(3 * kpn * N / R0 * (c[1] - c[0]))
            dc.append(
                3 / ((R0 - v * t) ** 3 - (R0 / 2) ** 3)
                * (kpn * (R0 / 2) ** 2 * (c[0] - c[1])
                   - kpn2 * (R0 - v * t) ** 2 * c[1])
            )
            return np.array(dc)

        def single_SPHERE(c, t):
            """This single ODE is solved if only one compartment remains"""
            # kpn = permeability coefficient (um/hour)
            kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t))
            if kpn2 >= 2 * D / (tolerance * (N - n + 1) * R0 / N):
                kpn2 = 2 * D / (tolerance * (N - n + 1) * R0 / N)
            dc = -3 * kpn2 / (R0 - v * t) * c[0]
            return np.array(dc)

    "###-=-=-=-=-=-=-=-=- Create Distribution -=-=-=-=-=-=-=-=-=-=-###"
    if use_distribution.get() is True:
        # Create a normal distribution with mean = R_0
        # Equally subdivide the continuum from min to max of the distribution
        n_bin1 = []  # Reassigns values in distribution to mid point of their bins
        n_bin = []  # Holds the number of particles inside each bin
        norm = np.random.normal(R0, SD, count)
        bound = np.max([R0 - np.min(norm), np.max(norm) - R0])
        x_bin = np.linspace(R0 - bound, R0 + bound, bins + 1)
        # For every bin, check to see if each particle is in that bin
        # If the particle is in that bin, reassign its value to the bins midpoint
        for j in range(len(x_bin) - 1):
            for i in range(count):
                if x_bin[j] <= norm[i] <= x_bin[j + 1]:
                    n_bin1.append((x_bin[j] + x_bin[j + 1]) / 2)
        # A loop that counts how many particles are in each bin
        # Stores values in n_bin
        for j in range(len(x_bin) - 1):
            n_bin2 = []
            for i in range(count):
                if n_bin1[i] == (x_bin[j] + x_bin[j + 1]) / 2:
                    n_bin2.append(1)
            n_bin.append(len(n_bin2))

    "###-=-=-=-=-=-=- If using distributions, solve inside each bin -=-=-=-=-=-###"
    if use_distribution.get() is True:
        # Solve the ODE in the q^th bin
        for q in range(bins):
            # Set the radius used in the ODEs to the midpoint of the q^th bin
            # (x_bin array holds the endpoints of each bin)
            R0 = (x_bin[q] + x_bin[q + 1]) / 2
            "###-=-=-=- Find at what time each compartment disappears -=-=-=-=-###"
            # Check to see when each of the i compartments disappears.
            # t_deg is an array that holds the degradation time points
            # Check every value of v*t.
            # The i^th compartment disappears when v*t becomes more than i*R0/n
            # The times of degradation are stored in the array t_deg
            t_deg = []
            for i in range(N + 1):
                for j in range(len(t)):
                    if v * t[j] >= i * R0 / N:
                        t_deg.append(j)
                        break
            # The final time point is added to allow the final system
            # of equations to be solved
            if len(t_deg) <= N:
                t_deg.append(len(t))

            "###-=-=-=-=-=-=- Solving the System in a Cylinder -=-=-=-=-=-=-###"
            if geometry == 'cylinder':
                "Cylinder-specific Parameters"
                vol = np.pi * R0 ** 2 * h  # Initial volume of polymer
                cm = (mass / vol)  # Initial concentration in the membrane (mg/mL)
                cb = (mass_bound / vol)  # Concentration of drug bound to the polymer
                c = np.ones(N) * cm  # Setup array with all initial conditions
                J1 = []  # Setup array to hold values for the flux

                "###-=-=-=-=-=-=-=-=- Run the ODE Solver -=-=-=-=-=-=-=-=-=-=-###"
                # Change the number of equations everytime a compartment disappears
                # n = the number of remaining compartments
                for i in range(len(t_deg) - 1):
                    n = N - i
                    # kpn2 = permeability coefficient in outside submembrane
                    kpn2 = 2 * D / ((N - n + 1) * R0 / N - v * t[t_deg[i]:t_deg[i + 1]])
                    # put a bound on kpn2
                    for j in range(len(kpn2)):
                        if kpn2[j] >= 2 * D / (tolerance * R0 / N):
                            kpn2[j] = 2 * D / (tolerance * R0 / N)

                    # If n>2 compartments remain, solve the system of n ODEs.
                    if n > 2:
                        solver = odeint(CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                        # J1 is the flux of free drug plus flux of bound drug.
                        # J1 = (kp+v)*C_n + v*C_b.
                        J1 = np.append(
                            J1,
                            solver[:, -1] * kpn2
                            + v * solver[:, -1] + v * cb
                        )
                        # Make current concentration the new initial condition.
                        c = []
                        for j in range(n - 1):
                            c = np.append(c, solver[-1, j])
                    # If 2 compartments remain, solve the system of 2 ODEs.
                    if n == 2:
                        solver = odeint(two_CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                        # Find J1 as before.
                        J1 = np.append(
                            J1,
                            solver[:, -1] * kpn2
                            + v * solver[:, -1] + v * cb
                        )
                        # Make current concentration the new initial condition.
                        c = []
                        for j in range(n - 1):
                            c = np.append(c, solver[-1, j])
                    # If 1 comparmtent remains, solve the single ODE.
                    if n == 1:
                        solver = odeint(single_CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                        # Find J1 as before.
                        J1 = np.append(
                            J1,
                            solver[:, -1] * kpn2
                            + v * solver[:, -1] + v * cb
                        )

                # J2 is the total flux
                # J2 = J1*(surface area)*(number of fibres in the q^th bin)
                # (n_bin[q] is the number of fibres in the q^th bin)
                J2 = 2 * np.pi * h * (R0 - v * t[0:t_deg[-1]]) * J1 * n_bin[q]
                # Q2 = integral(J2)
                Q2 = np.cumsum(J2 * (tm - t0) / ti)

            "###-=-=-=-=-=-=-=- Solving the System in a Sphere -=-=-=-=-=-=-=-###"
            if geometry == 'sphere':
                "Sphere-specific Parameters"
                vol = 4 * np.pi * R0 ** 3 / 3  # Initial volume of polymer
                cm = (mass / vol)  # Initial concentration in the membrane (mg/uL)
                cb = (mass_bound / vol)  # Concentration of drug bound to the polymer
                c = np.ones(N) * cm  # Setup array with all initial conditions
                J1 = []  # Setup array to hold values for the flux
                "###-=-=-=-=-=-=-=-=- Run the ODE Solver -=-=-=-=-=-=-=-=-=-=-###"
                # Change the number of equations everytime a compartment disappears
                # n = the number of remaining compartments
                for i in range(len(t_deg) - 1):
                    n = N - i
                    # kpn2 = permeability coefficient of the outside submembrane
                    kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t[t_deg[i]:t_deg[i + 1]]))
                    # put a bound on kpn2
                    for j in range(len(kpn2)):
                        if kpn2[j] >= 2 * D / (tolerance * R0 / N):
                            kpn2[j] = 2 * D / (tolerance * R0 / N)

                    # If more than 2 compartments remain, solve system of n ODEs.
                    if n > 2:
                        solver = odeint(SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                        # J1 is the flux of free drug plus flux of bound drug.
                        # J1 = (kp+v)*C_n + v*C_b.
                        J1 = np.append(
                            J1,
                            kpn2 * solver[:, -1]
                            + v * solver[:, -1] + v * cb
                        )
                        # Make current concentration the new initial condition.
                        c = []
                        for j in range(n - 1):
                            c = np.append(c, solver[-1, j])
                    # If 2 compartments remain, solve the system of 2 ODEs.
                    if n == 2:
                        solver = odeint(two_SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                        # Find J1 as before.
                        J1 = np.append(
                            J1,
                            kpn2 * solver[:, -1]
                            + v * solver[:, -1] + v * cb
                        )
                        # Make current concentration the new initial condition.
                        c = []
                        for j in range(n - 1):
                            c = np.append(c, solver[-1, j])
                    # If 1 comparmtent remains, solve the single ODE.
                    if n == 1:
                        solver = odeint(single_SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                        # Find J1 as before.
                        J1 = np.append(
                            J1,
                            kpn2 * solver[:, -1]
                            + v * solver[:, -1] + v * cb
                        )

                # J2 is the total flux
                # J2 = J1*(surface area)*(number of fibres in the q^th bin)
                # (n_bin[q] is the number of fibres in the q^th bin)
                J2 = 4 * np.pi * (R0 - v * t[0:t_deg[-1]]) ** 2 * J1 * n_bin[q]
                # Q2 = integral(J2)
                Q2 = np.cumsum(J2 * (tm - t0) / ti)

            # Add the values for flux or Q to the running total from previous bins
            for i in range(ti - len(Q2)):
                Q2 = np.append(Q2, Q2[-1])
            Q1 = Q1 + Q2
            # J_out is the flux
            for i in range(len(J1)):
                J_out[i] = J1[i] + J_out[i]
            for i in range(len(J1)):
                J_total[i] = J2[i] + J_total[i]
        J_out = J_out / (bins)

    "###-=-=-=-=-=- If not using distributions, solve with mean radius-=-=-=-=-###"
    if use_distribution.get() is False:

        "###-=-=-=-=- Find at what time each compartment disappears -=-=-=-=-=-###"
        # Check to see when each of the i compartments disappears.
        # t_deg is an array that holds the degradation time points
        # Check every value of v*t.
        # The i^th compartment disappears when v*t becomes more than i*R0/n
        # The times of degradation are stored in the array t_deg
        t_deg = []
        for i in range(N + 1):
            for j in range(len(t)):
                if v * t[j] >= i * R0 / N:
                    t_deg.append(j)
                    break
        # The final time point is added to allow the final system
        # of equations to be solved
        if len(t_deg) <= N:
            t_deg.append(len(t))

        "###-=-=-=-=-=-=-=- Solving the System in a Slab -=-=-=-=-=-=-=-###"
        if geometry == 'slab':
            "Slab-specific Parameters"
            area = slab_area  # Surface area
            vol = area * R0  # Volume of polymer
            cm = (mass / vol)  # Initial concentration in the membrane (ug/mL)
            cb = (mass_bound / vol)  # Concentration of drug bound to the polymer
            c = np.ones(N) * cm  # Setup array with all initial conditions
            J_out = []  # Setup array to hold values for the flux

            "###-=-=-=-=-=-=-=-=-=- Run the ODE Solver -=-=-=-=-=-=-=-=-=-=-=-###"
            # Change the number of equations everytime a compartment disappears
            # n = the number of remaining compartments
            for i in range(len(t_deg) - 1):
                n = N - i
                # kpn2 = the permeability coefficient of outside submembrane
                kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t[t_deg[i]:t_deg[i + 1]]))
                # stop kpn2 from approaching infinity
                for j in range(len(kpn2)):
                    if kpn2[j] >= 2 * D / (tolerance * R0 / N):
                        kpn2[j] = 2 * D / (tolerance * R0 / N)
                # If more than 2 compartments remain, solve the system of n ODEs.
                if n > 2:
                    solver = odeint(SLAB, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J_out, the flux of free drug plus flux of bound drug
                    # J_out = (kp+v)*C_n + v*C_b.
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 2 compartments remain, solve the system of 2 ODEs.
                if n == 2:
                    solver = odeint(two_SLAB, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J_out as before.
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 1 comparmtent remains, solve the single ODE.
                if n == 1:
                    solver = odeint(single_SLAB, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J_out as before
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
            # Calculate J_total, the total flux
            # J_total = J_out*(surface area)
            J_total = J_out * area
            # Calculate total amount of drug released.
            # Q1 = integral(J_total)
            Q1 = np.cumsum(J_total * (tm - t0) / ti)
            for i in range(ti - len(Q1)):
                Q1 = np.append(Q1, Q1[-1])

        "###-=-=-=-=-=-=-=- Solving the System in a Cylinder -=-=-=-=-=-=-=-###"
        if geometry == 'cylinder':
            "Cylinder-specific Parameters"
            vol = np.pi * R0 ** 2 * h  # Initial volume of polymer
            cm = (mass / vol)  # Initial concentration in the membrane (ug/mL)
            cb = (mass_bound / vol)  # Concentration of drug bound to the polymer
            c = np.ones(N) * cm  # Setup array with all initial conditions
            J_out = []  # Setup array to hold values for the flux

            "###-=-=-=-=-=-=-=-=-=- Run the ODE Solver -=-=-=-=-=-=-=-=-=-=-=-###"
            # Change the number of equations everytime a compartment disappears
            # n = the number of remaining compartments
            for i in range(len(t_deg) - 1):
                n = N - i
                # kpn2 = permeability coefficient of outside submembrane
                kpn2 = 2 * D / ((N - n + 1) * R0 / N - v * t[t_deg[i]:t_deg[i + 1]])
                # stop kpn2 from approaching infinity
                for j in range(len(kpn2)):
                    if kpn2[j] >= 2 * D / (tolerance * R0 / N):
                        kpn2[j] = 2 * D / (tolerance * R0 / N)

                # If n>2 compartments remain, solve the system of n ODEs.
                if n > 2:
                    solver = odeint(CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                    # J_out is the flux of free drug plus flux of bound drug.
                    # J_out = (kp+v)*C_n + v*C_b.
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 2 compartments remain, solve the system of 2 ODEs.
                if n == 2:
                    solver = odeint(two_CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J1 as before.
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 1 comparmtent remains, solve the single ODE.
                if n == 1:
                    solver = odeint(single_CYLINDER, c, t[t_deg[i]:t_deg[i + 1]])
                    # Calculate J_out as before
                    J_out = np.append(
                        J_out,
                        solver[:, -1] * kpn2
                        + v * solver[:, -1] + v * cb
                    )
            # Calculate J_total, the total flux out of all fibres
            # J_total = J_out*(surface area)
            J_total = 2 * np.pi * h * (R0 - v * t[0:t_deg[-1]]) * J_out * count
            # Q1 = integral(J_total).
            Q1 = np.cumsum(J_total * (tm - t0) / ti)
            for i in range(ti - len(Q1)):
                Q1 = np.append(Q1, Q1[-1])

        "###-=-=-=-=-=-=-=- Solving the System in a Sphere -=-=-=-=-=-=-=-###"
        if geometry == 'sphere':
            "Sphere-specific Parameters"
            vol = 4 * np.pi * R0 ** 3 / 3  # Initial volume of polymer
            cm = (mass / vol)  # Initial concentration in the membrane (mg/uL)
            cb = (mass_bound / vol)  # Concentration of drug bound to the polymer
            c = np.ones(N) * cm  # Setup array with all initial conditions
            J_out = []  # Setup array to hold values for the flux
            "###-=-=-=-=-=-=-=-=-=- Run the ODE Solver -=-=-=-=-=-=-=-=-=-=-=-###"
            # Change the number of equations everytime a compartment disappears
            # n = the number of remaining compartments
            for i in range(len(t_deg) - 1):
                n = N - i
                # kpn2 = permeability coefficient of outside submembrane
                kpn2 = 2 * D / (((N - n + 1) * R0 / N - v * t[t_deg[i]:t_deg[i + 1]]))
                # stop kpn2 from approaching infinity
                for j in range(len(kpn2)):
                    if kpn2[j] >= 2 * D / (tolerance * R0 / N):
                        kpn2[j] = 2 * D / (tolerance * R0 / N)

                # If more than 2 compartments remain, solve the system of n ODEs.
                if n > 2:
                    solver = odeint(SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                    # J1 is the flux of free drug plus flux of bound drug.
                    # J1 = (kp+v)*C_n + v*C_b.
                    J_out = np.append(
                        J_out,
                        kpn2 * solver[:, -1]
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 2 compartments remain, solve the system of 2 ODEs.
                if n == 2:
                    solver = odeint(two_SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J1 as before.
                    J_out = np.append(
                        J_out,
                        kpn2 * solver[:, -1]
                        + v * solver[:, -1] + v * cb
                    )
                    # Make the current concentration, the new initial condition.
                    c = []
                    for j in range(n - 1):
                        c = np.append(c, solver[-1, j])
                # If 1 comparmtent remains, solve the single ODE.
                if n == 1:
                    solver = odeint(single_SPHERE, c, t[t_deg[i]:t_deg[i + 1]])
                    # Find J1 as before.
                    J_out = np.append(
                        J_out,
                        kpn2 * solver[:, -1]
                        + v * solver[:, -1] + v * cb
                    )
            # Calculate J_total, the total flux out of all fibres
            # J_total = J_out*(surface area)
            J_total = 4 * np.pi * (R0 - v * t[0:t_deg[-1]]) ** 2 * J_out * count
            # Q1 = integral(J_total).
            Q1 = np.cumsum(J_total * (tm - t0) / ti)
            for i in range(ti - len(Q1)):
                Q1 = np.append(Q1, Q1[-1])

        "###-=-=-=-=-=-=-=-=-=- Calculations -=-=-=-=-=-=-=-=-###"

    if enter_experimental_data.get() is True:
        Q_out = clarify_Q(Q1, experimental_time, t)
        return Q_out


def saveSimulationPoints(xlist, ylist):
    fl = filedialog.asksaveasfile(mode='w', defaultextension=".txt")
    if fl is None:  # asksaveasfile return `None` if dialog closed with "cancel".
        return

    for x, y in zip(xlist, ylist):
        fl.write("%s\t%s\n" % (x, y))

    fl.close()


def clarify_Q(Q1, xm, t):
    # -=-=-=-=-=- Q_output is new to regression code. It holds the calculated release values at the experimental t values
    # -=-=-=-=-=- This is because the curve_fit program requires the input function to have a specific output of this type
    Q_output = []
    for j in xm[0]:
        for i in range(len(t)):
            if t[i] >= j:
                Q_output.append(Q1[i])
                break

    # -=-=-=-=-=- number_of_steps is new to regression code. It gets printed to screen as regression is calculated to show progress
    # number_of_steps += 1

    return Q_output


def define_input_parameters():
    Input_Parameters = []
    # Choose from pg, ng, ug, mg, g, kg
    if input_mass_var.get() == 1:
        Input_Parameters.append('pg')
    if input_mass_var.get() == 2:
        Input_Parameters.append('ng')
    if input_mass_var.get() == 3:
        Input_Parameters.append('ug')
    if input_mass_var.get() == 4:
        Input_Parameters.append('mg')
    if input_mass_var.get() == 5:
        Input_Parameters.append('g')
    if input_mass_var.get() == 6:
        Input_Parameters.append('kg')
    #             pm, nm, um, mm, cm, m
    if input_length_var.get() == 1:
        Input_Parameters.append('pm')
    if input_length_var.get() == 2:
        Input_Parameters.append('nm')
    if input_length_var.get() == 3:
        Input_Parameters.append('um')
    if input_length_var.get() == 4:
        Input_Parameters.append('mm')
    if input_length_var.get() == 5:
        Input_Parameters.append('cm')
    if input_length_var.get() == 6:
        Input_Parameters.append('m')
    #             s, min, hour, day, year
    if input_time_var.get() == 1:
        Input_Parameters.append('s')
    if input_time_var.get() == 2:
        Input_Parameters.append('min')
    if input_time_var.get() == 3:
        Input_Parameters.append('hour')
    if input_time_var.get() == 4:
        Input_Parameters.append('day')
    if input_time_var.get() == 5:
        Input_Parameters.append('year')
    return Input_Parameters


def define_output_parameters():
    Output_Parameters = []
    # Choose from pg, ng, ug, mg, g, kg
    if output_mass_var.get() == 1:
        Output_Parameters.append('pg')
    if output_mass_var.get() == 2:
        Output_Parameters.append('ng')
    if output_mass_var.get() == 3:
        Output_Parameters.append('ug')
    if output_mass_var.get() == 4:
        Output_Parameters.append('mg')
    if output_mass_var.get() == 5:
        Output_Parameters.append('g')
    if output_mass_var.get() == 6:
        Output_Parameters.append('kg')
    #             pm, nm, um, mm, cm, m
    if output_length_var.get() == 1:
        Output_Parameters.append('pm')
    if output_length_var.get() == 2:
        Output_Parameters.append('nm')
    if output_length_var.get() == 3:
        Output_Parameters.append('um')
    if output_length_var.get() == 4:
        Output_Parameters.append('mm')
    if output_length_var.get() == 5:
        Output_Parameters.append('cm')
    if output_length_var.get() == 6:
        Output_Parameters.append('m')
    #             s, min, hour, day, year
    if output_time_var.get() == 1:
        Output_Parameters.append('s')
    if output_time_var.get() == 2:
        Output_Parameters.append('min')
    if output_time_var.get() == 3:
        Output_Parameters.append('hour')
    if output_time_var.get() == 4:
        Output_Parameters.append('day')
    if output_time_var.get() == 5:
        Output_Parameters.append('year')
    return Output_Parameters


def get_entry_value(entr):
    vals = entr.get().split(' ')
    if len(vals) > 0:
        return float(vals[0])
    return float(entr.get())


def regression():
    if enter_experimental_data.get() is False:
        calculate(t, D, v)
    else:
        # # -=-=-=-=-=- iD and iv are initial guesses for diffusion coefficient and velocity of degradation
        # iD = 6.8870e-5
        # iv = 4.1186e-3
        # -=-=-=-=-=- curve_fit runs the regression, inputs are
        # -=-=-=-=-=- f is input function, xdata is an array of experimental time values, ym is an array of experimental release values
        # -=-=-=-=-=- p0 is an array containing the initial guess for diffusion coefficient, then the velocity of erosion
        # -=-=-=-=-=- bounds is a tuple of arrays, the first array is the lower bounds for D then v, the second array are upper bounds
        pars, cov = curve_fit(f=calculate, xdata=experimental_time[0], ydata=y_regression[0],
                              p0=[D, v],
                              bounds=([D / 100, v / 100], [D * 100, v * 100]),
                              )

        # -=-=-=-=-=- the following prints some relevant information to screen after calculations are complete
        global stat_D, stat_v, sd_D, sd_v
        stat_D = pars[0]
        stat_v = pars[1]
        sd_D = np.sqrt(np.diag(cov)[0])
        sd_v = np.sqrt(np.diag(cov)[1])


def save_calcilated_data():
    file_name = tk.filedialog.asksaveasfilename()

    wb = openpyxl.Workbook()
    i = 0
    t_label = 'Time ({})'.format(Output_Parameters[2])
    ws = wb.active

    # Plot the amount of drug released over time.
    if 'Q' in plot:
        if i == 0:
            ws.title = "Amount of Drug Released"
        else:
            ws = wb.create_sheet("Amount of Drug Released")

        Q_label = 'Amount of Drug Released ({})'.format(Output_Parameters[0])
        dataset = pd.DataFrame({t_label: t, Q_label: Q1}, columns=[t_label, Q_label])

        for r in dataframe_to_rows(dataset, index=False, header=True):
            ws.append(r)

        i = i + 1

    # # Plot the flux
    if 'flux' in plot:
        if i == 0:
            ws.title = "Flux per Unit Area"
        else:
            ws = wb.create_sheet("Flux per Unit Area")

        J_label = 'Flux per Unit Area ({}/{}/{}^2)'.format(Output_Parameters[0], Output_Parameters[2], Output_Parameters[1])
        dataset = pd.DataFrame({t_label: t[0:len(J_out)], J_label: J_out}, columns=[t_label, J_label])

        for r in dataframe_to_rows(dataset, index=False, header=True):
            ws.append(r)

        i = i + 1

    # Plot the concentration values
    if 'concentration' in plot:
        if i == 0:
            ws.title = "Flux per Unit Area"
        else:
            ws = wb.create_sheet("Concentration")

        R_label = 'Radius ({})'.format(Output_Parameters[1])
        C_label = 'Concentration ({}/{}^3)'.format(Output_Parameters[0], Output_Parameters[1])
        dataset = pd.DataFrame({R_label: np.linspace(R0 / N / 2,
                        R0 - (len(t_deg) - 2) * R0 / N - R0 / N / 2,
                        N - len(t_deg) + 2), C_label: solver[-1, :]},
                               columns=[R_label, C_label])

        for r in dataframe_to_rows(dataset, index=False, header=True):
            ws.append(r)

        i = i + 1

    # Plot the total flux
    if 'total_flux' in plot:
        if i == 0:
            ws.title = "Flux per Unit Area"
        else:
            ws = wb.create_sheet("Total Flux Out")

        F_label = 'Total Flux Out ({}/{})'.format(Output_Parameters[0], Output_Parameters[2])
        dataset = pd.DataFrame({t_label: t[0:len(J_total)], F_label: J_total}, columns=[t_label, F_label])

        for r in dataframe_to_rows(dataset, index=False, header=True):
            ws.append(r)

    wb.save(file_name + ".xlsx")


def show_results():
    rows = 2
    columns = 2
    x_text_size = 12
    y_text_size = 11
    Q_position = 3
    flux_position = 1
    concentration_position = 4
    total_flux_position = 2
    plt.subplots_adjust(
        wspace=0.5,
        hspace=0.5,
    )
    print("{:.2f}".format(Q1[-1]))
    release_label = tk.Label(output_results_root, text='RESULTS', font=arial)
    release_label.grid(row=0, column=0, padx=30, pady=15, sticky=tk.N)

    coefficients_label = tk.Label(output_results_root, text='Some input parameters for the graph are:', font=arial)
    coefficients_label.grid(row=1, column=0, padx=30, pady=3, sticky=tk.N)

    D_v_N_label = tk.Label(output_results_root,
                           text='D = {:.2e} {}^2/{},    v = {:.2e} {}/{},    N = {}'.format(D, Output_Parameters[1],
                                                                                            Output_Parameters[2],
                                                                                            v, Output_Parameters[1],
                                                                                            Output_Parameters[2], N),
                           font=arial)
    D_v_N_label.grid(row=2, column=0, padx=30, pady=3, sticky=tk.N)

    R_CV_label = tk.Label(output_results_root,
                          text='R0 = {:.2f} {},    CV = {:.2f}'.format(R0, Output_Parameters[1], SD / R0), font=arial)
    R_CV_label.grid(row=3, column=0, padx=30, pady=3, sticky=tk.N)

    mass_label = tk.Label(output_results_root, text='with {} of drug loaded and {} percent of drug boud'.format(mass,
                                                                                                                mass_bound / mass * 100),
                          font=arial)
    mass_label.grid(row=4, column=0, padx=30, pady=10, sticky=tk.N)

    # -=-=-=-=-=- the following prints some relevant information to screen after calculations are complete

    if enter_experimental_data.get() is True:
        info_best_fit_label = tk.Label(output_results_root, text="The parameters with best fit are below:", font=arial)
        info_best_fit_label.grid(row=5, column=0, padx=30, pady=3, sticky=tk.N)

        fit_D_label = tk.Label(output_results_root, text="D = {:.4e} {}^2/{} with standard deviation = {:.4e}".
                               format(stat_D, Output_Parameters[1], Output_Parameters[2], sd_D), font=arial)
        fit_D_label.grid(row=6, column=0, padx=30, pady=3, sticky=tk.N)

        fit_v_label = tk.Label(output_results_root, text="v = {:.4e} {}/{} with standard deviation = {:.4e}".
                               format(stat_v, Output_Parameters[1], Output_Parameters[2], sd_v), font=arial)
        fit_v_label.grid(row=7, column=0, padx=30, pady=3, sticky=tk.N)

    style.configure('Save.TButton', font=arial, borderless=1)
    save_calcilated_data_button = ttk.Button(output_results_root, text="Save fitted points",
                                             command=lambda: [save_calcilated_data()], style='Save.TButton')
    save_calcilated_data_button.grid(row=9, padx=30, column=0, pady=3)

    "###-=-=-=-=-=-=-=-=-=-=-=-=- Graph Results -=-=-=-=-=-=-=-=-=-=-=-=-=-###"
    if make_graphs.get() is True:
        # Plot experimental results
        if enter_experimental_data.get() is True:

            for i in range(len(experimental_flux)):
                if type(experimental_flux[i]) == np.ndarray:
                    if i == 0:
                        plt.subplot(rows, columns, flux_position)
                    plt.plot(experimental_time, experimental_flux[i])

            for i in range(len(experimental_total_flux)):
                if type(experimental_total_flux[i]) == np.ndarray:
                    if i == 0:
                        plt.subplot(rows, columns, total_flux_position)
                    plt.plot(experimental_time, experimental_total_flux[i])

            for i in range(len(experimental_release)):
                if type(experimental_release[i]) == np.ndarray:
                    if i == 0:
                        plt.subplot(rows, columns, Q_position)
                    plt.plot(experimental_time[0], experimental_release[i])

            for i in range(len(experimental_concentration)):
                if type(experimental_concentration[i]) == np.ndarray:
                    if i == 0:
                        plt.subplot(rows, columns, concentration_position)
                    plt.plot(experimental_x, experimental_concentration[i])

        # Plot the amount of drug released over time.
        if 'Q' in plot:
            plt.subplot(rows, columns, Q_position)
            plt.plot(t, Q1, ':b')
            plt.xlabel(f'Time (${Output_Parameters[2]}$)', fontsize=x_text_size)
            plt.ylabel(f'Amount of Drug Released (${Output_Parameters[0]}$)',
                       fontsize=y_text_size)

        saveSimulationPoints(t, Q1)

        # Plot the flux
        if 'flux' in plot:
            plt.subplot(rows, columns, flux_position)
            plt.plot(t[0:len(J_out)], J_out, 'r')
            plt.xlabel(f"Time (${Output_Parameters[2]}$)", fontsize=x_text_size)
            plt.ylabel(f"Flux per Unit Area "
                       f"(${Output_Parameters[0]}/{Output_Parameters[2]}"
                       f"/{Output_Parameters[1]}^2$)",
                       fontsize=y_text_size)

        # Plot the concentration values
        if 'concentration' in plot:
            plt.subplot(rows, columns, concentration_position)
            plt.plot(
                np.linspace(R0 / N / 2,
                            R0 - (len(t_deg) - 2) * R0 / N - R0 / N / 2,
                            N - len(t_deg) + 2),
                solver[-1, :],
                "o"
            )
            plt.xlabel(f"Radius (${Output_Parameters[1]}$)", fontsize=x_text_size)
            plt.ylabel("Concentration "
                       f"(${Output_Parameters[0]}/{Output_Parameters[1]}^3$)",
                       fontsize=y_text_size)
            plt.xticks(np.linspace(0, R0, N + 1))
            plt.ylim(0, np.max(c) * 1.1)

        # Plot the total flux
        if 'total_flux' in plot:
            plt.subplot(rows, columns, total_flux_position)
            plt.plot(t[0:len(J_total)], J_total, 'r')
            plt.xlabel(f"Time (${Output_Parameters[2]}$)",
                       fontsize=x_text_size)
            plt.ylabel("Total Flux Out "
                       f"(${Output_Parameters[0]}/{Output_Parameters[2]}$)",
                       fontsize=y_text_size)

        plt.suptitle("Results graphs", fontsize=y_text_size * 1.4)
        plt.show()


root = tk.Tk()
style = tk.ttk.Style()
root.title("Compartments")

width = (root.winfo_screenwidth() - root.winfo_reqwidth()) / 2
height = (root.winfo_screenheight() - root.winfo_reqheight()) / 4
# root.wm_geometry("+%d+%d" % (width/2, height))

arial = tk.font.Font(family='Times New Roman', size=12, weight=tk.font.BOLD)
info_font = tk.font.Font(family='Times New Roman', size=9, weight=tk.font.BOLD)

note = tk.ttk.Notebook(root)
note.pack(fill=tk.BOTH, expand=True)

tab01 = tk.ttk.Frame(note)
tab01.pack(fill=tk.BOTH, expand=True)

canvas=tk.Canvas(tab01)
vbar=tk.Scrollbar(tab01, orient=tk.VERTICAL, command=canvas.yview)
vbar.pack(side=tk.RIGHT, fill=tk.Y)
canvas.configure(yscrollcommand=vbar.set, scrollregion=canvas.bbox("all"))
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
tab1 = tk.Frame(canvas)
canvas.create_window((0,0), window=tab1, anchor=tk.NW)

tab2 = tk.ttk.Frame(note)
tab2.pack(fill=tk.BOTH, expand=True)
tab3 = tk.ttk.Frame(note)
tab3.pack(fill=tk.BOTH, expand=True)
note.add(tab01, text="Modeling")
note.add(tab2, text="Manual")
note.add(tab3, text="About")

######### TAB 1 Modeling ##########
"###-=-=-=-=-=-=-=-=-=- Choose what you want -=-=-=-=-=-=-=-=-=-=-###"

choose_label = tk.Label(tab1, text='Please, Choose what you want:', font=arial)
choose_label.grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
choose_root = tk.Label(tab1)
choose_root.grid(row=1, column=0, padx=10, sticky=tk.W)
enter_experimental_data = tk.BooleanVar()
enter_experimental_data.set(0)
enter_experimental_data_checkbox = tk.Checkbutton(choose_root, text="Enter experimental data",
                                                  variable=enter_experimental_data,
                                                  onvalue=1, offvalue=0)
enter_experimental_data_checkbox.pack(anchor=tk.W, padx=10)

use_distribution = tk.BooleanVar()
use_distribution.set(0)
use_distribution_checkbox = tk.Checkbutton(choose_root, text="Use distribution",
                                           variable=use_distribution,
                                           onvalue=1, offvalue=0)
use_distribution_checkbox.pack(anchor=tk.W, padx=10)

make_graphs = tk.BooleanVar()
make_graphs.set(0)
make_graphs_checkbox = tk.Checkbutton(choose_root, text="Make graphs",
                                      variable=make_graphs,
                                      onvalue=1, offvalue=0)
make_graphs_checkbox.pack(anchor=tk.W, padx=10)

"###-=-=-=-=-=-=- Choose the geometry -=-=-=-=-=-=-###"

geometry_label = tk.Label(tab1, text='Please, choose the geometry:', font=arial)
geometry_label.grid(row=2, column=1, padx=30, pady=10, sticky=tk.W)
geometry_root = tk.Label(tab1)
geometry_root.grid(row=3, column=1, padx=30, sticky=tk.W)
geometry_var = tk.IntVar()
geometry_var.set(0)
rbutton_slab = tk.Radiobutton(geometry_root, text='slab', variable=geometry_var, value=1,
                              command=lambda: [choose_geometry(1)])
rbutton_cylinder = tk.Radiobutton(geometry_root, text='cylinder', variable=geometry_var, value=2,
                                  command=lambda: [choose_geometry(2)])
rbutton_sphere = tk.Radiobutton(geometry_root, text='sphere', variable=geometry_var, value=3,
                                command=lambda: [choose_geometry(3)])
rbutton_slab.grid(row=0, column=0, padx=10, sticky=tk.W)
rbutton_cylinder.grid(row=0, column=1, padx=10, sticky=tk.W)
rbutton_sphere.grid(row=0, column=2, padx=10, sticky=tk.W)


def choose_geometry(val):
    geometry_var.set(val)


"###-=-=-=-=-=-=- Choose what to plot -=-=-=-=-=-=-###"

plot_label = tk.Label(tab1, text='Please, Choose what to plot:', font=arial)
plot_label.grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
plot_root = tk.Label(tab1)
plot_root.grid(row=3, column=0, padx=10, sticky=tk.W)
plot_q = tk.BooleanVar()
plot_q.set(0)
plot_q_checkbox = tk.Checkbutton(plot_root, text="total release",
                                 variable=plot_q,
                                 onvalue=1, offvalue=0)
plot_q_checkbox.grid(row=0, column=0, padx=5, sticky=tk.W)
plot_flux = tk.BooleanVar()
plot_flux.set(0)
plot_flux_checkbox = tk.Checkbutton(plot_root, text="flux",
                                    variable=plot_flux,
                                    onvalue=1, offvalue=0)
plot_flux_checkbox.grid(row=0, column=1, padx=5, sticky=tk.W)
plot_total_flux = tk.BooleanVar()
plot_total_flux.set(0)
plot_total_flux_checkbox = tk.Checkbutton(plot_root, text="total flux",
                                          variable=plot_total_flux,
                                          onvalue=1, offvalue=0)
plot_total_flux_checkbox.grid(row=1, column=0, padx=5, sticky=tk.W)
plot_concentration = tk.BooleanVar()
plot_concentration.set(0)
plot_concentration_checkbox = tk.Checkbutton(plot_root, text="concentration",
                                             variable=plot_concentration,
                                             onvalue=1, offvalue=0)
plot_concentration_checkbox.grid(row=1, column=1, padx=5, sticky=tk.W)

"###-=-=-=-=-=-=-=- Define Parameters and Constants -=-=-=-=-=-=-=-###"
input_parameters_label = tk.Label(tab1, text='Decide the units of your input parameters:', font=arial)
input_parameters_label.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky=tk.W)
input_parameters_root = tk.Label(tab1)
input_parameters_root.grid(row=5, column=0, columnspan=3, padx=10, sticky=tk.W)

# Choose from pg, ng, ug, mg, g, kg

mass_label = tk.Label(input_parameters_root, text='Mass')
mass_label.grid(row=0, column=0, padx=10, sticky=tk.W)

input_mass_var = tk.IntVar()
input_mass_var.set(0)
intput_pg = tk.Radiobutton(input_parameters_root, text='pg', variable=input_mass_var, value=1,
                           command=lambda: [choose_input_mass(1)])
intput_pg.grid(row=1, column=0, padx=10, sticky=tk.W)
intput_ng = tk.Radiobutton(input_parameters_root, text='ng', variable=input_mass_var, value=2,
                           command=lambda: [choose_input_mass(2)])
intput_ng.grid(row=2, column=0, padx=10, sticky=tk.W)
intput_ug = tk.Radiobutton(input_parameters_root, text='ug', variable=input_mass_var, value=3,
                           command=lambda: [choose_input_mass(3)])
intput_ug.grid(row=3, column=0, padx=10, sticky=tk.W)
intput_mg = tk.Radiobutton(input_parameters_root, text='mg', variable=input_mass_var, value=4,
                           command=lambda: [choose_input_mass(4)])
intput_mg.grid(row=4, column=0, padx=10, sticky=tk.W)
intput_g = tk.Radiobutton(input_parameters_root, text='g', variable=input_mass_var, value=5,
                          command=lambda: [choose_input_mass(5)])
intput_g.grid(row=5, column=0, padx=10, sticky=tk.W)
intput_kg = tk.Radiobutton(input_parameters_root, text='kg', variable=input_mass_var, value=6,
                           command=lambda: [choose_input_mass(6)])
intput_kg.grid(row=6, column=0, padx=10, sticky=tk.W)

# intput_percent = tk.Radiobutton(input_parameters_root, text='%', variable=input_mass_var, value=7,
#                            command=lambda: [choose_input_mass(7)])
# intput_percent.grid(row=7, column=0, padx=10, sticky=tk.W)


def choose_input_mass(val):
    input_mass_var.set(val)


#             pm, nm, um, mm, cm, m

length_label = tk.Label(input_parameters_root, text='Length')
length_label.grid(row=0, column=1, padx=10, sticky=tk.W)
input_length_var = tk.IntVar()
input_length_var.set(0)
intput_pm = tk.Radiobutton(input_parameters_root, text='pm', variable=input_length_var, value=1,
                           command=lambda: [choose_input_length(1)])
intput_pm.grid(row=1, column=1, padx=10, sticky=tk.W)
intput_nm = tk.Radiobutton(input_parameters_root, text='nm', variable=input_length_var, value=2,
                           command=lambda: [choose_input_length(2)])
intput_nm.grid(row=2, column=1, padx=10, sticky=tk.W)
intput_um = tk.Radiobutton(input_parameters_root, text='um', variable=input_length_var, value=3,
                           command=lambda: [choose_input_length(3)])
intput_um.grid(row=3, column=1, padx=10, sticky=tk.W)
intput_mm = tk.Radiobutton(input_parameters_root, text='mm', variable=input_length_var, value=4,
                           command=lambda: [choose_input_length(4)])
intput_mm.grid(row=4, column=1, padx=10, sticky=tk.W)
intput_cm = tk.Radiobutton(input_parameters_root, text='cm', variable=input_length_var, value=5,
                           command=lambda: [choose_input_length(5)])
intput_cm.grid(row=5, column=1, padx=10, sticky=tk.W)
intput_m = tk.Radiobutton(input_parameters_root, text='m', variable=input_length_var, value=6,
                          command=lambda: [choose_input_length(6)])
intput_m.grid(row=6, column=1, padx=10, sticky=tk.W)


def choose_input_length(val):
    input_length_var.set(val)


#             s, min, hour, day, year

time_label = tk.Label(input_parameters_root, text='Time')
time_label.grid(row=0, column=2, padx=10, sticky=tk.W)
input_time_var = tk.IntVar()
input_time_var.set(0)
intput_s = tk.Radiobutton(input_parameters_root, text='s', variable=input_time_var, value=1,
                          command=lambda: [choose_input_time(1)])
intput_s.grid(row=1, column=2, padx=10, sticky=tk.W)
intput_min = tk.Radiobutton(input_parameters_root, text='min', variable=input_time_var, value=2,
                            command=lambda: [choose_input_time(2)])
intput_min.grid(row=2, column=2, padx=10, sticky=tk.W)
intput_hour = tk.Radiobutton(input_parameters_root, text='hour', variable=input_time_var, value=3,
                             command=lambda: [choose_input_time(3)])
intput_hour.grid(row=3, column=2, padx=10, sticky=tk.W)
intput_day = tk.Radiobutton(input_parameters_root, text='day', variable=input_time_var, value=4,
                            command=lambda: [choose_input_time(4)])
intput_day.grid(row=4, column=2, padx=10, sticky=tk.W)
intput_year = tk.Radiobutton(input_parameters_root, text='year', variable=input_time_var, value=5,
                             command=lambda: [choose_input_time(5)])
intput_year.grid(row=5, column=2, padx=10, sticky=tk.W)


def choose_input_time(val):
    input_time_var.set(val)


### OUTPUT PARAMETERS ####
output_parameters_label = tk.Label(tab1, text='Decide the units to graph results with:', font=arial)
output_parameters_label.grid(row=4, column=1, columnspan=3, padx=30, pady=10, sticky=tk.W)
output_parameters_root = tk.Label(tab1)
output_parameters_root.grid(row=5, column=1, columnspan=3, padx=30, sticky=tk.W)

# Choose from pg, ng, ug, mg, g, kg

mass_label = tk.Label(output_parameters_root, text='Mass')
mass_label.grid(row=0, column=0, padx=10, sticky=tk.W)

output_mass_var = tk.IntVar()
output_mass_var.set(0)
intput_pg = tk.Radiobutton(output_parameters_root, text='pg', variable=output_mass_var, value=1,
                           command=lambda: [choose_output_mass(1)])
intput_pg.grid(row=1, column=0, padx=10, sticky=tk.W)
intput_ng = tk.Radiobutton(output_parameters_root, text='ng', variable=output_mass_var, value=2,
                           command=lambda: [choose_output_mass(2)])
intput_ng.grid(row=2, column=0, padx=10, sticky=tk.W)
intput_ug = tk.Radiobutton(output_parameters_root, text='ug', variable=output_mass_var, value=3,
                           command=lambda: [choose_output_mass(3)])
intput_ug.grid(row=3, column=0, padx=10, sticky=tk.W)
intput_mg = tk.Radiobutton(output_parameters_root, text='mg', variable=output_mass_var, value=4,
                           command=lambda: [choose_output_mass(4)])
intput_mg.grid(row=4, column=0, padx=10, sticky=tk.W)
intput_g = tk.Radiobutton(output_parameters_root, text='g', variable=output_mass_var, value=5,
                          command=lambda: [choose_output_mass(5)])
intput_g.grid(row=5, column=0, padx=10, sticky=tk.W)
intput_kg = tk.Radiobutton(output_parameters_root, text='kg', variable=output_mass_var, value=6,
                           command=lambda: [choose_output_mass(6)])
intput_kg.grid(row=6, column=0, padx=10, sticky=tk.W)


def choose_output_mass(val):
    output_mass_var.set(val)


#             pm, nm, um, mm, cm, m

length_label = tk.Label(output_parameters_root, text='Length')
length_label.grid(row=0, column=1, padx=10, sticky=tk.W)
output_length_var = tk.IntVar()
output_length_var.set(0)
intput_pm = tk.Radiobutton(output_parameters_root, text='pm', variable=output_length_var, value=1,
                           command=lambda: [choose_output_length(1)])
intput_pm.grid(row=1, column=1, padx=10, sticky=tk.W)
intput_nm = tk.Radiobutton(output_parameters_root, text='nm', variable=output_length_var, value=2,
                           command=lambda: [choose_output_length(2)])
intput_nm.grid(row=2, column=1, padx=10, sticky=tk.W)
intput_um = tk.Radiobutton(output_parameters_root, text='um', variable=output_length_var, value=3,
                           command=lambda: [choose_output_length(3)])
intput_um.grid(row=3, column=1, padx=10, sticky=tk.W)
intput_mm = tk.Radiobutton(output_parameters_root, text='mm', variable=output_length_var, value=4,
                           command=lambda: [choose_output_length(4)])
intput_mm.grid(row=4, column=1, padx=10, sticky=tk.W)
intput_cm = tk.Radiobutton(output_parameters_root, text='cm', variable=output_length_var, value=5,
                           command=lambda: [choose_output_length(5)])
intput_cm.grid(row=5, column=1, padx=10, sticky=tk.W)
intput_m = tk.Radiobutton(output_parameters_root, text='m', variable=output_length_var, value=6,
                          command=lambda: [choose_output_length(6)])
intput_m.grid(row=6, column=1, padx=10, sticky=tk.W)


def choose_output_length(val):
    output_length_var.set(val)


#             s, min, hour, day, year

time_label = tk.Label(output_parameters_root, text='Time')
time_label.grid(row=0, column=2, padx=10, sticky=tk.W)
output_time_var = tk.IntVar()
output_time_var.set(0)
intput_s = tk.Radiobutton(output_parameters_root, text='s', variable=output_time_var, value=1,
                          command=lambda: [choose_output_time(1)])
intput_s.grid(row=1, column=2, padx=10, sticky=tk.W)
intput_min = tk.Radiobutton(output_parameters_root, text='min', variable=output_time_var, value=2,
                            command=lambda: [choose_output_time(2)])
intput_min.grid(row=2, column=2, padx=10, sticky=tk.W)
intput_hour = tk.Radiobutton(output_parameters_root, text='hour', variable=output_time_var, value=3,
                             command=lambda: [choose_output_time(3)])
intput_hour.grid(row=3, column=2, padx=10, sticky=tk.W)
intput_day = tk.Radiobutton(output_parameters_root, text='day', variable=output_time_var, value=4,
                            command=lambda: [choose_output_time(4)])
intput_day.grid(row=4, column=2, padx=10, sticky=tk.W)
intput_year = tk.Radiobutton(output_parameters_root, text='year', variable=output_time_var, value=5,
                             command=lambda: [choose_output_time(5)])
intput_year.grid(row=5, column=2, padx=10, sticky=tk.W)


def choose_output_time(val):
    output_time_var.set(val)


######## Other Parameters #########
# N = 10  # Number of compartments
# count = 1000  # Number of particles / fibres
# mass = 120  # Mass of drug loaded (M)
# mass_bound = 0  # Mass of drug bound to polymer (M)
# v = 1.9e-8  # Erosion front velocity (L/T)
# D = 5e-9  # Diffusion coefficient (L^2/T)
# R0 = 0.1  # Initial radius/width (L)
# h = 1000  # Height of cylinder (L)
# slab_area = 1000  # Area if solving in a slab (L^2)
# kp = N * D / h  # Permeability coefficient with constant volume (L/T)
# t0 = 0  # Start time (T)
# tm = R0 ** 2 / D  # End time (T)
# ti = 10000  # How many time intervals
# tolerance = 0.01  # Determine when to stop kpn2 from growing
# SD = R0 / 100  # Standard deviation of distribution (L)
# bins = 10  # Number of bins in distribution

emply_label = tk.Label(tab1, text='')
emply_label.grid(row=6, column=0, sticky=tk.S)

N_label = tk.Label(tab1, text='Number of compartments:')
N_label.grid(row=8, column=0, padx=15, pady=3, sticky=tk.E)
entry_N = tk.Entry(tab1, width=15)
entry_N.grid(row=8, column=1, padx=15, pady=3, sticky=tk.W)
entry_N.insert(0, '10 (default)')

ti_label = tk.Label(tab1, text='Number of time intervals:')
ti_label.grid(row=9, column=0, padx=15, pady=3, sticky=tk.E)
entry_ti = tk.Entry(tab1, width=15)
entry_ti.grid(row=9, column=1, padx=15, pady=3, sticky=tk.W)
entry_ti.insert(0, '10000 (default)')

count_label = tk.Label(tab1, text='Number of particles / fibres:')
count_label.grid(row=10, column=0, padx=15, pady=3, sticky=tk.E)
entry_count = tk.Entry(tab1, width=15)
entry_count.grid(row=10, column=1, padx=15, pady=3, sticky=tk.W)
entry_count.insert(0, '1000 (default)')

mass_label = tk.Label(tab1, text='Mass loaded in each particle / fibre [M]:')
mass_label.grid(row=11, column=0, padx=15, pady=3, sticky=tk.E)
entry_mass = tk.Entry(tab1, width=15)
entry_mass.grid(row=11, column=1, padx=15, pady=3, sticky=tk.W)
entry_mass.insert(0, '120 (default)')

mass_bound_label = tk.Label(tab1, text='Mass bound to each particle / fibre [M]:')
mass_bound_label.grid(row=12, column=0, padx=15, pady=3, sticky=tk.E)
entry_mass_bound = tk.Entry(tab1, width=15)
entry_mass_bound.grid(row=12, column=1, padx=15, pady=3, sticky=tk.W)
entry_mass_bound.insert(0, '0 (default)')

v_label = tk.Label(tab1, text='Erosion front velocity [L/T]:')
v_label.grid(row=13, column=0, padx=15, pady=3, sticky=tk.E)
entry_v = tk.Entry(tab1, width=15)
entry_v.grid(row=13, column=1, padx=15, pady=3, sticky=tk.W)
entry_v.insert(0, '1.9e-8 (default)')

D_label = tk.Label(tab1, text='Diffusion coefficient [L^2/T]:')
D_label.grid(row=14, column=0, padx=15, pady=3, sticky=tk.E)
entry_D = tk.Entry(tab1, width=15)
entry_D.grid(row=14, column=1, padx=15, pady=3, sticky=tk.W)
entry_D.insert(0, '5e-9 (default)')

R0_label = tk.Label(tab1, text='Initial radius/width [L]:')
R0_label.grid(row=15, column=0, padx=15, pady=3, sticky=tk.E)
entry_R0 = tk.Entry(tab1, width=15)
entry_R0.grid(row=15, column=1, padx=15, pady=3, sticky=tk.W)
entry_R0.insert(0, '0.1 (default)')

h_label = tk.Label(tab1, text='Height of cylinder [L]:')
h_label.grid(row=16, column=0, padx=15, pady=3, sticky=tk.E)
entry_h = tk.Entry(tab1, width=15)
entry_h.grid(row=16, column=1, padx=15, pady=3, sticky=tk.W)
entry_h.insert(0, '1000 (default)')

slab_area_label = tk.Label(tab1, text='Area of slab [L^2]:')
slab_area_label.grid(row=17, column=0, padx=15, pady=3, sticky=tk.E)
entry_slab_area = tk.Entry(tab1, width=15)
entry_slab_area.grid(row=17, column=1, padx=15, pady=3, sticky=tk.W)
entry_slab_area.insert(0, '1000 (default)')

t0_label = tk.Label(tab1, text='Start time [T]:')
t0_label.grid(row=18, column=0, padx=15, pady=3, sticky=tk.E)
entry_t0 = tk.Entry(tab1, width=15)
entry_t0.grid(row=18, column=1, padx=15, pady=3, sticky=tk.W)
entry_t0.insert(0, '0 (default)')

tm_label = tk.Label(tab1, text='End time [T]:')
tm_label.grid(row=19, column=0, padx=15, pady=3, sticky=tk.E)
entry_tm = tk.Entry(tab1, width=15)
entry_tm.grid(row=19, column=1, padx=15, pady=3, sticky=tk.W)
entry_tm.insert(0, '100 (default)')

tolerance_label = tk.Label(tab1, text='Tolerance:')
tolerance_label.grid(row=20, column=0, padx=15, pady=3, sticky=tk.E)
entry_tolerance = tk.Entry(tab1, width=15)
entry_tolerance.grid(row=20, column=1, padx=15, pady=3, sticky=tk.W)
entry_tolerance.insert(0, '0.01 (default)')

SD_label = tk.Label(tab1, text='Standard deviation [L]:')
SD_label.grid(row=21, column=0, padx=15, pady=3, sticky=tk.E)
entry_SD = tk.Entry(tab1, width=15)
entry_SD.grid(row=21, column=1, padx=15, pady=3, sticky=tk.W)
entry_SD.insert(0, '0 (default)')

bins_label = tk.Label(tab1, text='Number of bins in distribution:')
bins_label.grid(row=22, column=0, padx=15, pady=3, sticky=tk.E)
entry_bins = tk.Entry(tab1, width=15)
entry_bins.grid(row=22, column=1, padx=15, pady=3, sticky=tk.W)
entry_bins.insert(0, '10 (default)')

"###-=-=-=-=-=-=-=- Experimental data -=-=-=-=-=-=-=-###"


def read_experimental_data():
    file_name = tk.filedialog.askopenfilename()
    # import data as numpy arrays
    global imported_data_1
    if type(file_name) is str:
        try:
            imported_data_1 = np.array(pd.read_excel(file_name, header=None, engine='openpyxl')).T
        except Exception:
            tk.messagebox.showerror(title="invalid file format!", message="openpyxl does not support the old .xls "
                                                                          "file format, please use xlrd to read this "
                                                                          "file, or convert it to the more recent "
                                                                          ".xlsx file format.")


style.configure('Open.TButton', font=arial, borderless=1)
data_file_button = ttk.Button(tab1, text="Open experimental data file (.xlsx)",
                              command=lambda: [read_experimental_data()], style='Open.TButton')
data_file_button.grid(row=8, column=2, pady=3, columnspan=2)

columns_label = tk.Label(tab1, text='Please enter column numbers for your\nexperimental data (actual only, starting at '
                                    '1)', font=arial)
columns_label.grid(row=10, rowspan=2, column=2, padx=10, pady=3, sticky=tk.N, columnspan=2)

time_column_label = tk.Label(tab1, text='Time column number:')
time_column_label.grid(row=12, column=2, padx=10, pady=3, sticky=tk.W)
entry_time_column = tk.Entry(tab1, width=10)
entry_time_column.grid(row=12, column=3, padx=10, pady=3, sticky=tk.W)

release_column_label = tk.Label(tab1, text='Release column number:')
release_column_label.grid(row=13, column=2, padx=10, pady=3, sticky=tk.W)
entry_release_column = tk.Entry(tab1, width=10)
entry_release_column.grid(row=13, column=3, padx=10, pady=3, sticky=tk.W)

flux_column_label = tk.Label(tab1, text='Flux column number:')
flux_column_label.grid(row=14, column=2, padx=10, pady=3, sticky=tk.W)
entry_flux_column = tk.Entry(tab1, width=10)
entry_flux_column.grid(row=14, column=3, padx=10, pady=3, sticky=tk.W)

total_flux_column_label = tk.Label(tab1, text='Total flux column number:')
total_flux_column_label.grid(row=15, column=2, padx=10, pady=3, sticky=tk.W)
entry_total_flux_column = tk.Entry(tab1, width=10)
entry_total_flux_column.grid(row=15, column=3, padx=10, pady=3, sticky=tk.W)

x_column_label = tk.Label(tab1, text='Flux column number:')
x_column_label.grid(row=16, column=2, padx=10, pady=3, sticky=tk.W)
entry_x_column = tk.Entry(tab1, width=10)
entry_x_column.grid(row=16, column=3, padx=10, pady=3, sticky=tk.W)

concentration_column_label = tk.Label(tab1, text='Flux column number:')
concentration_column_label.grid(row=17, column=2, padx=10, pady=3, sticky=tk.W)
entry_concentration_column = tk.Entry(tab1, width=10)
entry_concentration_column.grid(row=17, column=3, padx=10, pady=3, sticky=tk.W)

"###-=-=-=-=-=-=-=- Output results -=-=-=-=-=-=-=-###"

output_results_root = tk.Label(tab1, highlightthickness=1)
output_results_root.grid(row=0, column=2, pady=10, sticky=tk.N, columnspan=2, rowspan=5)

######### TAB 2 GUIDE ##########

info_label = tk.Label(tab2, text='Step-by-step Instructions', font=arial). \
    grid(row=0, column=0, padx=10, sticky=tk.N, pady=10)
info1_label = tk.Label(tab2, text='(1) Under "choose what you want" section, decide whether you have ' +
                                  'experimental data, or a known distribution and if you want to plot ' +
                                  'graphs by changing the truth value of the appropriate variable.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=1, column=0, padx=10, sticky=tk.W, pady=2)
info2_label = tk.Label(tab2, text='(2) Under the "choose geometry" section, enter the geometry of the ' +
                                  'problem, which results you want to plot and some style choices for the plots.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=2, column=0, padx=10, sticky=tk.W, pady=2)
info3_label = tk.Label(tab2, text='(3) Under the "define parameters" section, enter the units of your input ' +
                                  'parameters and the units that you want to be graphed (output), then ' +
                                  'enter all parameters.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=3, column=0, padx=10, sticky=tk.W, pady=2)
info4_label = tk.Label(tab2, text='(4) If you have experimental data. Under "Enter experimental data" ' +
                                  'section, import your data from an excel document (up to 5 files), it ' +
                                  'will be converted into a numpy array with transposed rows/columns. ' +
                                  'Then assign the data to either flux, total flux, concentration or ' +
                                  'release, (up to 5 of each) as appropriate.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=4, column=0, padx=10, sticky=tk.W, pady=2)

info5_label = tk.Label(tab2, text='Notes About Parameter Choices', font=arial). \
    grid(row=5, column=0, padx=10, sticky=tk.N, pady=10)
info6_label = tk.Label(tab2, text='- Don\'t make mass and mass_bound too small.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=6, column=0, padx=10, sticky=tk.W, pady=2)
info7_label = tk.Label(tab2, text='- The tolerance should be much smaller for slabs than other geometries.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=7, column=0, padx=10, sticky=tk.W, pady=2)
info8_label = tk.Label(tab2, text='- The larger the value chosen for N is, ti must chosen sufficiently large and ' +
                                  'tolerance sufficiently low.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=8, column=0, padx=10, sticky=tk.W, pady=2)
info9_label = tk.Label(tab2, text='- Make sure SD is appropriately small for the choice of R0, if SD is too large ' +
                                  'there will be particles with negative radius, the equality on line 287 ' +
                                  'will always hold and the solver will not run, and an error will be ' +
                                  'produced on line 395.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=9, column=0, padx=10, sticky=tk.W, pady=2)
nfo10_label = tk.Label(tab2, text='- When v is made too large, discontinuities in the flux can become bad ' +
                                  'v should not be more than 1 order of magnitude larger than D.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=10, column=0, padx=10, sticky=tk.W, pady=2)
nfo11_label = tk.Label(tab2, text='- Diffusion coefficients are much smaller in nanoparticles/fibres than in slabs.',
                       wraplength=root.winfo_reqwidth() * 5).grid(row=11, column=0, padx=10, sticky=tk.W, pady=2)

######### TAB 3 ABOUT ##########

about_label = tk.Label(tab3, text="This program based on the compartmental and size distribution models " +
                                  "and was developed by the Griffith University and " +
                                  "Tomsk Polytechnic University research team \n " +
                                  "(Anissimov Y.G., Spiridonova T.I., Marriott R. and Petlin D.G.)",
                       wraplength=root.winfo_reqwidth() * 2).pack(pady=10)
about1_label = tk.Label(tab3,
                        text='The contact e-mails:\ny.anissimov@griffith.edu.au\nrory.marriott@griffithuni.edu.au',
                        wraplength=root.winfo_reqwidth() * 2, anchor='w').pack(pady=10)

################################

style.configure('Calculate.TButton', background='#3CB371', font=arial, width=20)
b5 = ttk.Button(tab1, text='Calculate', command=lambda: [check(), define_params(), regression(), show_results()], style='Calculate.TButton')
b5.grid(row=21, column=2, sticky=tk.E)

emply_label = tk.Label(tab1, text='')
emply_label.grid(row=22, column=0)

root.mainloop()

# Changes Rory has made: 
#   Edited the wording of some of the parameter labels
#   Added a display of some important parameters
#   Changed the output text to only display after a calculation is performed
